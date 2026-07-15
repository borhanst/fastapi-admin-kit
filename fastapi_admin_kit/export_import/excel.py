"""Excel Export/Import implementation (optional, requires openpyxl)."""

from __future__ import annotations

import io
from typing import Any

from fastapi_admin_kit.export_import.base import ExportBase, ImportBase

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class ExcelExport(ExportBase):
    """Export data to Excel format (.xlsx).

    Requires openpyxl to be installed: pip install openpyxl

    Usage::

        class ProductExport(ExcelExport):
            columns = ["id", "name", "price"]
            headers = {"name": "Product Name", "price": "Price (USD)"}
    """

    def __init__(self, registered: Any) -> None:
        if not HAS_OPENPYXL:
            raise ImportError(
                "Excel export requires openpyxl. Install it with: pip install openpyxl"
            )
        super().__init__(registered)

    def export(self, queryset: Any, request: Any = None) -> io.BytesIO:
        """Export queryset to Excel format.

        Args:
            queryset: SQLAlchemy queryset or list of objects
            request: Optional request object

        Returns:
            BytesIO object containing Excel data
        """
        columns = self.get_columns()
        headers = self.get_headers()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self.registered.verbose_name_plural or "Export"

        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0EA5E9", end_color="0EA5E9", fill_type="solid")

        # Write headers
        for col_idx, col in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=headers.get(col, col))
            cell.font = header_font
            cell.fill = header_fill

        # Write data
        row_count = 0
        if queryset is not None:
            for obj in queryset:
                if self.max_rows and row_count >= self.max_rows:
                    break
                row_count += 1
                for col_idx, col in enumerate(columns, 1):
                    value = self.get_value(obj, col)
                    value = self.format_value(value, col)
                    ws.cell(row=row_count + 1, column=col_idx, value=value)

        # Auto-adjust column widths
        for col_idx, col in enumerate(columns, 1):
            max_length = len(headers.get(col, col))
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(
                max_length + 2, 50
            )

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output


class ExcelImport(ImportBase):
    """Import data from Excel format (.xlsx).

    Requires openpyxl to be installed: pip install openpyxl

    Usage::

        class ProductImport(ExcelImport):
            required_columns = ["name", "price"]
            field_map = {"Product Name": "name", "Price": "price"}
            unique_key = "id"
    """

    def __init__(self, registered: Any) -> None:
        if not HAS_OPENPYXL:
            raise ImportError(
                "Excel import requires openpyxl. Install it with: pip install openpyxl"
            )
        super().__init__(registered)

    def parse(self, file_content: bytes, request: Any = None) -> list[dict[str, Any]]:
        """Parse Excel file content and return list of row dictionaries.

        Args:
            file_content: Raw Excel file content as bytes
            request: Optional request object

        Returns:
            List of dictionaries, each representing a row
        """
        wb = openpyxl.load_workbook(io.BytesIO(file_content))
        ws = wb.active

        # Get headers from first row
        headers = []
        for cell in ws[1]:
            headers.append(str(cell.value).strip() if cell.value else "")

        # Parse rows
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {}
            for idx, value in enumerate(row):
                if idx < len(headers) and headers[idx]:
                    # Convert None to empty string
                    row_dict[headers[idx]] = value if value is not None else ""
            rows.append(row_dict)

        return rows
